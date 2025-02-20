# Excel library methods
# Excel manager class (used by report engine)
# utility classes for common formatting for use by reports
#

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill


class CellFormatTitle(object):
    # Styling for table headers
    font = Font(bold=True, size=16)
    fill = PatternFill()
    alignment = Alignment(horizontal="left", vertical="center")
    border = Border(
        left=Side(), right=Side(),
        top=Side(), bottom=Side(style="thick")
    )


class CellFormatHeader(object):
    # Styling for table headers
    font = Font(bold=True, color="FFFFFF")
    fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    alignment = Alignment(horizontal="center", vertical="center")
    border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )


class CellFormatBody(object):
    # Styling for table body
    font = Font()
    fill = PatternFill()
    alignment = Alignment(horizontal="center", vertical="center")
    border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )


class CellFormatFixed(object):
    # Styling for Fixed Width Text (raw text - unknown length)
    font = Font(name="Courier New", size=12)
    fill = PatternFill()
    alignment = Alignment()
    border = Border()


class ExcelManager:
    def __init__(self):
        self.workbook = None
        self.file_path = None

    # ###############################################################
    # Workbook management methods
    #

    def create_spreadsheet(self, file_path):
        """
        Creates a new Excel workbook and initializes it.
        """
        self.workbook = Workbook()
        self.file_path = file_path
        # Optionally remove the default sheet created by openpyxl
        default_sheet = self.workbook.active
        self.workbook.remove(default_sheet)
        print(f"Created a new spreadsheet: {file_path}")

    def open_existing_spreadsheet(self, file_path):
        """
        Opens an existing Excel workbook for editing.
        """
        self.workbook = load_workbook(file_path)
        self.file_path = file_path
        print(f"Opened existing spreadsheet: {file_path}")

    def save(self):
        """
        Saves the current workbook.
        """
        if not self.workbook:
            raise Exception("No active workbook to save.")
        self.workbook.save(self.file_path)

    def close(self):
        """
        Saves the current workbook.
        """
        if not self.workbook:
            raise Exception("No active workbook to close.")
        self.workbook.close()
        self.workbook = None

    def save_and_close(self):
        """
        Saves and closes the current workbook.
        """
        if not self.workbook:
            raise Exception("No active workbook to save.")
        self.workbook.save(self.file_path)
        self.workbook.close()
        self.workbook = None
        print(f"Saved and closed the spreadsheet: {self.file_path}")

    # ###############################################################
    # Generic methods to create new tabs
    #
    # need methods to hanld different ways data could come in.
    #
    #   Text processor
    #   List processor
    #   dictionary processor

    def add_tab_by_string(self, tab_name, str_output, title=None,
                          title_format=CellFormatTitle(),
                          body_format=CellFormatFixed()):

        """
        Adds a new tab to the workbook and write raw text data to it.
        """
        if not self.workbook:
            raise Exception("No active workbook. Create or open a spreadsheet first.")
        sheet = self.workbook.create_sheet(title=tab_name)

        # TODO: Error check for it tab_name already exists
        # NOTE: Unique formatting on this one so far

        if title:
            cell = sheet.cell(row=1, column=1, value=title)
            cell.font = title_format.font
            cell.fill = title_format.fill
            cell.alignment = title_format.alignment
            cell.border = title_format.border

        lines = str_output.split('\n')
        for row_idx, row in enumerate(lines, start=2):
            cell = sheet.cell(row=row_idx, column=1, value=row)
            cell.font = body_format.font
            cell.fill = body_format.fill
            cell.alignment = body_format.alignment
            cell.border = body_format.border

        print(f"Added a new tab '{tab_name}' using straight TEXT to the workbook.")

    def add_tab_by_list(self, tab_name, raw_list_data, title='My Title'):
        """
        Adds a new tab to the workbook and write raw text data to it.
        """
        if not self.workbook:
            raise Exception("No active workbook. Create or open a spreadsheet first.")
        sheet = self.workbook.create_sheet(title=tab_name)

        # TODO: Error check for it tab_name already exists

        if title:
            cell = sheet.cell(row=1, column=1, value=title)

        for row_idx, row in enumerate(raw_list_data, start=2):
            cell = sheet.cell(row=row_idx, column=1, value=row)

        print(f"Added a new tab '{tab_name}' with formatted data to the workbook.")

    def add_tab_with_data(self, tab_name, data_dict):
        """
        Adds a new tab to the workbook and populates it with data from a dictionary.
        Keys of the dictionary become headers, and values are written row-wise.

        :param tab_name:    string of the tab_name
        :param: data_dict   DICTIONARY of data to write
        """
        if not self.workbook:
            raise Exception("No active workbook. Create or open a spreadsheet first.")

        # Create a new sheet
        sheet = self.workbook.create_sheet(title=tab_name)

        # Write headers
        headers = list(data_dict.keys())
        sheet.append(headers)

        # Write data (assuming values are lists of equal length)
        rows = zip(*data_dict.values())
        for row in rows:
            sheet.append(row)
        print(f"Added a new tab '{tab_name}' to the workbook.")

    # ENHANCEMENTS
    #   Add a method for a 2 or 3 level level header

    def add_tab_with_formatted_data(self, tab_name, data_dict,
                                    header_format=CellFormatHeader(),
                                    body_format=CellFormatBody()):
        """Add a new tab as a simple formatted table.

        Formatting objects are taken from defaults but can be overwriten.

        TODO: Finalize data format for receiving data

        :param tab_name:    new tab name
        :param data_dict:   properly formatted dictionary of data to write
        :param header_format:
        :param body_foramt:

        :return:    Nothing.  Updates excel workbook object in self
        """

        """
        Data format example:
            sample_data = {
                "Name": ["Alice", "Bob", "Charlie"],
                "Age": [25, 30, 35],
                "City": ["New York", "Los Angeles", "Chicago"]
            }

        Keys are Column Headers
        List attached is row data
        """

        if not self.workbook:
            raise Exception("No active workbook. Create or open a spreadsheet first.")
        sheet = self.workbook.create_sheet(title=tab_name)

        # TODO: Error check for it tab_name already exists

        # Write headers with formatting
        headers = list(data_dict.keys())
        for col_num, header in enumerate(headers, start=1):
            cell = sheet.cell(row=1, column=col_num, value=header)
            cell.font = header_format.font
            cell.fill = header_format.fill
            cell.alignment = header_format.alignment
            cell.border = header_format.border

        # Write data with formatting
        rows = zip(*data_dict.values())
        for row_idx, row in enumerate(rows, start=2):  # Start from second row
            for col_idx, value in enumerate(row, start=1):
                cell = sheet.cell(row=row_idx, column=col_idx, value=value)
                cell.font = body_format.font
                cell.fill = body_format.fill
                cell.alignment = body_format.alignment
                cell.border = body_format.border

        # Auto-adjust column widths
        for col_num, col_cells in enumerate(sheet.columns, start=1):
            max_length = max(len(str(cell.value or "")) for cell in col_cells)
            sheet.column_dimensions[sheet.cell(row=1, column=col_num).column_letter].width = max_length + 2

        print(f"Added a new tab '{tab_name}' with formatted data to the workbook.")


# Example Usage for testing
if __name__ == "__main__":
    manager = ExcelManager()

    # Create and initialize a spreadsheet
    spreadsheet_name = 'example.xlsx'
    manager.create_spreadsheet(spreadsheet_name)

    # Add a new tab with data
    sample_data = {
        "Name": ["Alice", "Bob", "Charlie"],
        "Age": [25, 30, 35],
        "City": ["New York", "Los Angeles", "Chicago"]
    }

    # new tab without formatting
    manager.add_tab_with_data("People", sample_data)

    # new tab with sample formatting
    manager.add_tab_with_formatted_data("People2", sample_data)

    # Save and close
    manager.save_and_close()

    if False:
        # Reopen the existing spreadsheet and add another tab
        manager.open_existing_spreadsheet(spreadsheet_name)
        another_data = {
            "Product": ["Laptop", "Mouse", "Keyboard"],
            "Price": [1000, 50, 70],
            "Stock": [10, 200, 150]
        }
        manager.add_tab_with_data("Inventory", another_data)
        manager.save_and_close()
